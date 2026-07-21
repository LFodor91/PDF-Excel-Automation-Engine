import re
import time
import pickle
from pathlib import Path
from typing import Dict, List, Optional
from collections import defaultdict

import pdfplumber
import openpyxl
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter
from watchdog.observers import Observer
from watchdog.events import FileSystemEventHandler

from PIL import Image, ImageOps, ImageFilter
import pytesseract


# =========================
# PATHS
# =========================
BASE_DIR = Path(__file__).resolve().parent

INPUT_DIR = BASE_DIR / "Input"           # GE / SAJATKESZLET / UNIMAS
INPUT_EK_DIR = BASE_DIR / "Input_EK"     # EK (PDF + kép)
OUTPUT_DIR = BASE_DIR / "Output"

LOOKUP_FILE = BASE_DIR / "lookup.xlsx"
CUSTOMERS_FILE = BASE_DIR / "Vevők.xlsx"

TEMPLATES_DIR = BASE_DIR / "Templates"
if not TEMPLATES_DIR.exists():
    TEMPLATES_DIR = BASE_DIR / "templates"

EK_TEMPLATE_FILE = TEMPLATES_DIR / "EK_template.xlsx"
GE_TEMPLATE_FILE = TEMPLATES_DIR / "GE_template.xlsx"
SK_TEMPLATE_FILE = TEMPLATES_DIR / "SAJATKESZLET_template.xlsx"
UNIMAS_TEMPLATE_FILE = TEMPLATES_DIR / "UNIMAS_template.xlsx"

TESSERACT_EXE = BASE_DIR / "Tesseract-OCR" / "tesseract.exe"

CACHE_DIR = BASE_DIR / "Cache"
LOOKUP_CACHE_FILE = CACHE_DIR / "lookup_cache.pkl"
CUSTOMERS_CACHE_FILE = CACHE_DIR / "customers_cache.pkl"
CACHE_VERSION = 3
ENGINE_VERSION = "2026.06-cache-v4-huf-format"

# EUR vevőnél a HUF EK ár ennyivel lesz elosztva.
EUR_RATE = 400.0
DEFAULT_CURRENCY = "HUF"

FILL_WARN = PatternFill(
    start_color="FFF2CC",
    end_color="FFF2CC",
    fill_type="solid",
)


# =========================
# PERSISTENT CACHE
# =========================
_EK_LOOKUP_CACHE: Optional[Dict[str, dict]] = None
_GE_LOOKUP_CACHE: Optional[Dict[str, str]] = None
_CUSTOMER_CURRENCY_CACHE: Optional[Dict[str, str]] = None


def _source_signature(source_path: Path) -> dict:
    stat = source_path.stat()
    return {"size": stat.st_size, "mtime_ns": stat.st_mtime_ns}


def _read_persistent_cache(cache_path: Path, source_path: Path):
    if not cache_path.exists():
        print(f"[CACHE] Nincs még cache: {cache_path.name}", flush=True)
        return None

    try:
        print(f"[CACHE] Ellenőrzés: {cache_path.name}", flush=True)
        with cache_path.open("rb") as file:
            payload = pickle.load(file)

        if payload.get("version") != CACHE_VERSION:
            print(f"[CACHE] Elavult cache: {cache_path.name}", flush=True)
            return None

        if payload.get("source_signature") != _source_signature(source_path):
            print(f"[CACHE] A forrás Excel módosult: {source_path.name}", flush=True)
            return None

        print(f"[CACHE] Betöltve: {cache_path.name}", flush=True)
        return payload.get("data")
    except Exception as error:
        print(f"[CACHE][WARN] Nem olvasható {cache_path.name}: {error}", flush=True)
        return None


def _write_persistent_cache(cache_path: Path, source_path: Path, data) -> None:
    CACHE_DIR.mkdir(parents=True, exist_ok=True)
    payload = {
        "version": CACHE_VERSION,
        "source_signature": _source_signature(source_path),
        "data": data,
    }
    temp_path = cache_path.with_suffix(cache_path.suffix + ".tmp")
    print(f"[CACHE] Mentés: {cache_path.name}", flush=True)
    with temp_path.open("wb") as file:
        pickle.dump(payload, file, protocol=pickle.HIGHEST_PROTOCOL)
    temp_path.replace(cache_path)
    print(f"[CACHE] Elkészült: {cache_path.name}", flush=True)


# =========================
# ROBUSTNESS / SPEED
# =========================
def wait_until_file_ready(
    path: Path,
    timeout: int = 30,
    step: float = 0.25,
) -> None:
    start = time.time()
    last_size = -1
    stable_count = 0

    while time.time() - start < timeout:
        try:
            size = path.stat().st_size
        except FileNotFoundError:
            time.sleep(step)
            continue

        if size > 0 and size == last_size:
            stable_count += 1
            if stable_count >= 2:
                return
        else:
            stable_count = 0
            last_size = size

        time.sleep(step)


def extract_text_from_pdf_pages(pdf_path: Path, max_pages: int) -> str:
    parts: List[str] = []

    with pdfplumber.open(str(pdf_path)) as pdf:
        for i, page in enumerate(pdf.pages):
            if i >= max_pages:
                break

            parts.append(page.extract_text() or "")

    return "\n".join(parts)


def extract_text_from_pdf_all(pdf_path: Path) -> str:
    parts: List[str] = []

    with pdfplumber.open(str(pdf_path)) as pdf:
        for page in pdf.pages:
            parts.append(page.extract_text() or "")

    return "\n".join(parts)


# =========================
# TEMPLATE HELPERS
# =========================
def safe_headers_from_template(ws) -> List[str]:
    headers: List[str] = []

    for c in range(1, ws.max_column + 1):
        value = ws.cell(1, c).value

        if isinstance(value, str) and value.strip():
            headers.append(value.strip())

    return headers


def shrink_first_excel_table(ws, last_col_letter: str) -> None:
    last_row = ws.max_row

    if last_row < 1:
        return

    new_ref = f"A1:{last_col_letter}{last_row}"

    table_obj = None

    if hasattr(ws, "tables") and ws.tables:
        table_obj = list(ws.tables.values())[0]
    elif hasattr(ws, "_tables") and ws._tables:
        table_obj = list(ws._tables.values())[0]

    if table_obj is not None:
        table_obj.ref = new_ref

        if getattr(table_obj, "autoFilter", None) is not None:
            table_obj.autoFilter.ref = new_ref


def resize_template_table(ws, header_count: int) -> None:
    if header_count < 1:
        return

    shrink_first_excel_table(
        ws,
        get_column_letter(header_count),
    )


# =========================
# COMMON
# =========================
def norm_key(value: str) -> str:
    text = "" if value is None else str(value).strip().upper()
    return re.sub(r"[^0-9A-Z]", "", text)


def parse_float_any(value) -> float:
    text = str(value).strip().replace("\u00a0", "").replace(" ", "")

    if text.startswith(".") or text.startswith(","):
        text = "0" + text

    # 1.234,56 -> 1234.56
    if "," in text and "." in text:
        if text.rfind(",") > text.rfind("."):
            text = text.replace(".", "").replace(",", ".")
        else:
            text = text.replace(",", "")
    else:
        text = text.replace(",", ".")

    return float(text)


def safe_float(value) -> Optional[float]:
    if value is None or value == "":
        return None

    if isinstance(value, (int, float)):
        return float(value)

    try:
        return parse_float_any(value)
    except (TypeError, ValueError):
        return None


# ============================================================
# VEVŐI PÉNZNEM
# ============================================================
def normalize_customer_number(value) -> str:
    """
    Példák:
      0000666734 -> 666734
      666734     -> 666734
      666734.0   -> 666734
    """
    if value is None:
        return ""

    if isinstance(value, float) and value.is_integer():
        text = str(int(value))
    else:
        text = str(value).strip()

    if text.endswith(".0"):
        text = text[:-2]

    digits = re.sub(r"\D", "", text)

    if not digits:
        return ""

    normalized = digits.lstrip("0")
    return normalized or "0"


def load_customer_currency_map() -> Dict[str, str]:
    global _CUSTOMER_CURRENCY_CACHE

    if _CUSTOMER_CURRENCY_CACHE is not None:
        print("[CACHE] Vevői pénznemek már memóriában.", flush=True)
        return _CUSTOMER_CURRENCY_CACHE

    if not CUSTOMERS_FILE.exists():
        print(
            f"[FIGYELEM] Hiányzik {CUSTOMERS_FILE.name}; minden EK ár HUF marad.",
            flush=True,
        )
        _CUSTOMER_CURRENCY_CACHE = {}
        return _CUSTOMER_CURRENCY_CACHE

    cached = _read_persistent_cache(CUSTOMERS_CACHE_FILE, CUSTOMERS_FILE)
    if cached is not None:
        _CUSTOMER_CURRENCY_CACHE = cached
        print(
            f"[CACHE] Vevői pénznemek kész: {len(_CUSTOMER_CURRENCY_CACHE)} vevő",
            flush=True,
        )
        return _CUSTOMER_CURRENCY_CACHE

    print(f"[EXCEL] Megnyitás: {CUSTOMERS_FILE.name}", flush=True)
    wb = load_workbook(CUSTOMERS_FILE, data_only=True, read_only=True)
    ws = wb.active
    rows = ws.iter_rows(values_only=True)
    first_row = next(rows, None)

    if not first_row:
        wb.close()
        raise ValueError("A Vevők.xlsx üres.")

    headers: Dict[str, int] = {}
    for idx, value in enumerate(first_row):
        if isinstance(value, str) and value.strip():
            headers[value.strip().lower()] = idx

    if "vevőszám" not in headers or "pénznem" not in headers:
        wb.close()
        raise ValueError(
            "A Vevők.xlsx első sorában kell a 'Vevőszám' és a 'Pénznem' fejléc."
        )

    customer_col = headers["vevőszám"]
    currency_col = headers["pénznem"]
    currency_map: Dict[str, str] = {}
    processed = 0

    print("[EXCEL] Vevők.xlsx beolvasása...", flush=True)
    for row in rows:
        processed += 1
        customer_value = row[customer_col] if customer_col < len(row) else None
        currency_value = row[currency_col] if currency_col < len(row) else None
        customer_number = normalize_customer_number(customer_value)

        if customer_number:
            currency = (
                str(currency_value).strip().upper()
                if currency_value is not None
                else DEFAULT_CURRENCY
            )
            if currency not in {"HUF", "EUR"}:
                currency = DEFAULT_CURRENCY
            currency_map[customer_number] = currency

        if processed % 5000 == 0:
            print(f"[EXCEL] Vevők.xlsx: {processed} sor...", flush=True)

    wb.close()
    print(
        f"[EXCEL] Vevők.xlsx kész: {processed} sor, {len(currency_map)} vevő",
        flush=True,
    )
    _write_persistent_cache(CUSTOMERS_CACHE_FILE, CUSTOMERS_FILE, currency_map)
    _CUSTOMER_CURRENCY_CACHE = currency_map
    return _CUSTOMER_CURRENCY_CACHE

def get_customer_currency(
    customer_number: str,
    customer_currency_map: Dict[str, str],
) -> str:
    normalized = normalize_customer_number(customer_number)

    return customer_currency_map.get(
        normalized,
        DEFAULT_CURRENCY,
    )


def convert_ek_price(price, currency: str):
    """
    HUF:
      egész számként kerül az Excelbe, tizedesjel nélkül.

    EUR:
      HUF EK / 400, két tizedesre kerekítve.
      Példa: 1208 / 400 = 3,02.
    """
    numeric_price = safe_float(price)

    if numeric_price is None:
        return "" if price is None else price

    if currency.upper() == "EUR":
        return round(numeric_price / EUR_RATE, 2)

    return int(round(numeric_price))


# ============================================================
# EK / SAJATKESZLET LOOKUP
# ============================================================
EK_LOOKUP_SHEET_NAME = "Sheet1"
EK_LOOKUP_COL_CODE = 2  # B
EK_LOOKUP_COL_NAME = 4  # D
EK_LOOKUP_COL_AE = 5    # E
EK_LOOKUP_COL_EK = 8    # H


def ek_normalize_code(value) -> str:
    if value is None:
        return ""

    tokens = re.findall(r"\d+", str(value))

    if not tokens:
        return ""

    # 6-3-3
    if (
        len(tokens) >= 3
        and len(tokens[0]) == 6
        and len(tokens[1]) == 3
        and len(tokens[2]) == 3
    ):
        return "".join(tokens[:3])

    # >=7-3-4
    if (
        len(tokens) >= 3
        and len(tokens[0]) >= 7
        and len(tokens[1]) == 3
        and len(tokens[2]) == 4
    ):
        return "".join(tokens[:3])

    # >=7-3-2/3
    if (
        len(tokens) >= 3
        and len(tokens[0]) >= 7
        and len(tokens[1]) == 3
        and len(tokens[2]) in (2, 3)
    ):
        return "".join(tokens[:3])

    # 6-2-3-2
    if (
        len(tokens) >= 4
        and len(tokens[0]) == 6
        and len(tokens[1]) == 2
        and len(tokens[2]) == 3
        and len(tokens[3]) == 2
    ):
        return "".join(tokens[:4])

    # 6-2-3-1
    # Példa: 107414 34 000 1
    if (
        len(tokens) >= 4
        and len(tokens[0]) == 6
        and len(tokens[1]) == 2
        and len(tokens[2]) == 3
        and len(tokens[3]) == 1
    ):
        return "".join(tokens[:4])

    # 5/6-(1 vagy 2)-3-4
    if (
        len(tokens) >= 4
        and len(tokens[0]) in (5, 6)
        and len(tokens[1]) in (1, 2)
        and len(tokens[2]) == 3
        and len(tokens[3]) == 4
    ):
        return "".join(tokens[:4])

    # 5/6-(1 vagy 2)-3-3
    if (
        len(tokens) >= 4
        and len(tokens[0]) in (5, 6)
        and len(tokens[1]) in (1, 2)
        and len(tokens[2]) == 3
        and len(tokens[3]) == 3
    ):
        return "".join(tokens[:4])

    # 5-3-4
    if (
        len(tokens) >= 3
        and len(tokens[0]) == 5
        and len(tokens[1]) == 3
        and len(tokens[2]) == 4
    ):
        return "".join(tokens[:3])

    # 5-3
    if (
        len(tokens) >= 2
        and len(tokens[0]) == 5
        and len(tokens[1]) == 3
    ):
        return "".join(tokens[:2])

    out: List[str] = []

    for token in tokens:
        if len(token) <= 2 and out:
            break

        out.append(token)

    return "".join(out)


def _load_lookup_maps_with_cache() -> tuple[Dict[str, dict], Dict[str, str]]:
    global _EK_LOOKUP_CACHE, _GE_LOOKUP_CACHE

    if _EK_LOOKUP_CACHE is not None and _GE_LOOKUP_CACHE is not None:
        print("[CACHE] Lookup már memóriában.", flush=True)
        return _EK_LOOKUP_CACHE, _GE_LOOKUP_CACHE

    if not LOOKUP_FILE.exists():
        raise FileNotFoundError(f"Nem találom a lookup fájlt: {LOOKUP_FILE}")

    cached = _read_persistent_cache(LOOKUP_CACHE_FILE, LOOKUP_FILE)
    if cached is not None:
        _EK_LOOKUP_CACHE = cached.get("ek", {})
        _GE_LOOKUP_CACHE = cached.get("ge", {})
        print(
            f"[CACHE] Lookup kész: EK/SK={len(_EK_LOOKUP_CACHE)}, GE={len(_GE_LOOKUP_CACHE)}",
            flush=True,
        )
        return _EK_LOOKUP_CACHE, _GE_LOOKUP_CACHE

    print(f"[EXCEL] Megnyitás: {LOOKUP_FILE.name}", flush=True)
    wb = openpyxl.load_workbook(LOOKUP_FILE, data_only=True, read_only=True)

    ek_ws = wb[EK_LOOKUP_SHEET_NAME] if EK_LOOKUP_SHEET_NAME in wb.sheetnames else wb.active
    ek_mapping: Dict[str, dict] = {}
    processed = 0
    print(f"[EXCEL] EK/SK lap: {ek_ws.title}", flush=True)

    for row in ek_ws.iter_rows(min_row=3, values_only=True):
        processed += 1
        sap = row[EK_LOOKUP_COL_CODE - 1] if len(row) >= EK_LOOKUP_COL_CODE else None
        name_hu = row[EK_LOOKUP_COL_NAME - 1] if len(row) >= EK_LOOKUP_COL_NAME else None
        ae = row[EK_LOOKUP_COL_AE - 1] if len(row) >= EK_LOOKUP_COL_AE else None
        ek = row[EK_LOOKUP_COL_EK - 1] if len(row) >= EK_LOOKUP_COL_EK else None
        key = ek_normalize_code(sap)

        if key:
            ek_mapping[key] = {
                "code": str(sap).strip() if sap is not None else "",
                "name": str(name_hu).strip() if name_hu is not None else "",
                "AE": ae,
                "EK": ek,
            }

        if processed % 5000 == 0:
            print(f"[EXCEL] EK/SK: {processed} sor...", flush=True)

    print(
        f"[EXCEL] EK/SK kész: {processed} sor, {len(ek_mapping)} kulcs",
        flush=True,
    )

    if GE_LOOKUP_SHEET_NAME in wb.sheetnames:
        ge_ws = wb[GE_LOOKUP_SHEET_NAME]
    elif len(wb.worksheets) >= 2:
        ge_ws = wb.worksheets[1]
    else:
        wb.close()
        raise ValueError("A lookup.xlsx fájlban nincs Sheet2 / második GE munkalap.")

    ge_rows = ge_ws.iter_rows(values_only=True)
    first_row = next(ge_rows, None)
    if not first_row:
        wb.close()
        raise ValueError("A GE lookup munkalap üres.")

    headers: Dict[str, int] = {}
    for idx, value in enumerate(first_row):
        if isinstance(value, str) and value.strip():
            headers[value.strip()] = idx

    if GE_LOOKUP_COL_ITEM not in headers or GE_LOOKUP_COL_RECA not in headers:
        wb.close()
        raise ValueError(
            f"A GE lookup fejléc hibás. Kell: '{GE_LOOKUP_COL_ITEM}' és "
            f"'{GE_LOOKUP_COL_RECA}'. Talált: {list(headers.keys())}"
        )

    col_item = headers[GE_LOOKUP_COL_ITEM]
    col_reca = headers[GE_LOOKUP_COL_RECA]
    ge_mapping: Dict[str, str] = {}
    ge_processed = 0
    print(f"[EXCEL] GE lap: {ge_ws.title}", flush=True)

    for row in ge_rows:
        ge_processed += 1
        item_value = row[col_item] if col_item < len(row) else None
        reca_value = row[col_reca] if col_reca < len(row) else None
        if item_value is not None:
            item_text = str(item_value).strip()
            if item_text:
                reca_text = "" if reca_value is None else str(reca_value).strip()
                ge_mapping[norm_key(item_text)] = reca_text

        if ge_processed % 5000 == 0:
            print(f"[EXCEL] GE: {ge_processed} sor...", flush=True)

    wb.close()
    print(
        f"[EXCEL] GE kész: {ge_processed} sor, {len(ge_mapping)} kulcs",
        flush=True,
    )

    data = {"ek": ek_mapping, "ge": ge_mapping}
    _write_persistent_cache(LOOKUP_CACHE_FILE, LOOKUP_FILE, data)
    _EK_LOOKUP_CACHE = ek_mapping
    _GE_LOOKUP_CACHE = ge_mapping
    return _EK_LOOKUP_CACHE, _GE_LOOKUP_CACHE


def ek_load_lookup_map() -> Dict[str, dict]:
    ek_map, _ = _load_lookup_maps_with_cache()
    return ek_map

def build_lookup_lengths(
    lookup_map: Dict[str, dict],
) -> List[int]:
    return sorted(
        {len(key) for key in lookup_map.keys()},
        reverse=True,
    )


def best_lookup_candidate_from_tokens(
    tokens: List[str],
    lookup_map: Dict[str, dict],
) -> str:
    if not tokens:
        return ""

    candidates: List[str] = []
    accumulated = ""

    for token in tokens:
        accumulated += token
        candidates.append(accumulated)

    for candidate in sorted(
        set(candidates),
        key=len,
        reverse=True,
    ):
        if candidate in lookup_map:
            return candidate

    return ""


# ============================================================
# EK PDF TEXT PARSER
# ============================================================
def ek_extract_customer_no(pdf_text: str) -> str:
    match = re.search(
        r"Vevőszám:\s*0*(\d+)",
        pdf_text,
        flags=re.IGNORECASE,
    )

    return match.group(1) if match else "UNKNOWN"


def ek_is_item_start_line(line: str) -> bool:
    match = re.match(r"^(\d{1,3})\s+(.*)$", line)

    if not match:
        return False

    after_position = match.group(2).strip()
    tokens = re.findall(r"\d+", after_position)

    if not tokens:
        return False

    if (
        len(tokens) >= 3
        and len(tokens[0]) == 6
        and len(tokens[1]) == 3
        and len(tokens[2]) == 3
    ):
        return True

    if (
        len(tokens) >= 3
        and len(tokens[0]) >= 7
        and len(tokens[1]) == 3
        and len(tokens[2]) == 4
    ):
        return True

    if (
        len(tokens) >= 3
        and len(tokens[0]) >= 7
        and len(tokens[1]) == 3
        and len(tokens[2]) in (2, 3)
    ):
        return True

    if (
        len(tokens) >= 4
        and len(tokens[0]) == 6
        and len(tokens[1]) == 2
        and len(tokens[2]) == 3
        and len(tokens[3]) in (1, 2)
    ):
        return True

    if (
        len(tokens) >= 4
        and len(tokens[0]) in (5, 6)
        and len(tokens[1]) in (1, 2)
        and len(tokens[2]) == 3
        and len(tokens[3]) in (3, 4)
    ):
        return True

    if (
        len(tokens) >= 3
        and len(tokens[0]) == 5
        and len(tokens[1]) == 3
        and len(tokens[2]) == 4
    ):
        return True

    if (
        len(tokens) >= 2
        and len(tokens[0]) == 5
        and len(tokens[1]) == 3
    ):
        return True

    if len(tokens[0]) >= 6:
        return True

    return False


def ek_extract_code_from_line(
    line: str,
    lookup_map: Dict[str, dict],
) -> str:
    match = re.match(r"^\d{1,3}\s+(.*)$", line)

    if not match:
        return ""

    rest = match.group(1).strip()
    tokens = re.findall(r"\d+", rest)

    if not tokens:
        return ""

    candidates: List[str] = []
    accumulated = ""

    for token in tokens:
        accumulated += token
        candidates.append(accumulated)

    for candidate in sorted(
        set(candidates),
        key=len,
        reverse=True,
    ):
        if candidate in lookup_map:
            return candidate

    return ek_normalize_code(rest)


def ek_parse_pdf_items(
    pdf_text: str,
    lookup_map: Dict[str, dict],
) -> List[dict]:
    lines = [
        line.strip()
        for line in pdf_text.splitlines()
        if line.strip()
    ]
    items: List[dict] = []

    for line in lines:
        low = line.lower()

        if (
            low.startswith("poz.")
            or "vevőszám" in low
            or low.startswith("megrendelés")
        ):
            continue

        if not ek_is_item_start_line(line):
            continue

        key = ek_extract_code_from_line(
            line,
            lookup_map,
        )

        if not key:
            continue

        items.append({"key": key})

    return items


# =========================
# EK OUTPUT
# =========================
def ek_write_output(
    items: List[dict],
    out_path: Path,
    lookup_map: Dict[str, dict],
    currency: str = DEFAULT_CURRENCY,
) -> tuple[int, int]:
    wb = load_workbook(EK_TEMPLATE_FILE)
    ws = wb.active

    if ws.max_row >= 2:
        ws.delete_rows(
            2,
            ws.max_row - 1,
        )

    headers = safe_headers_from_template(ws)

    if not headers:
        raise ValueError(
            "EK_template.xlsx első sora üres."
        )

    currency = (
        currency.upper()
        if currency
        else DEFAULT_CURRENCY
    )

    if currency not in {"HUF", "EUR"}:
        currency = DEFAULT_CURRENCY

    missing = 0

    for item in items:
        hit = lookup_map.get(item["key"])
        row_data: Dict[str, object] = {}

        if hit:
            row_data["Cikkszám"] = hit["code"]
            row_data["Megnevezés"] = hit["name"]
            row_data["EK"] = convert_ek_price(
                hit["EK"],
                currency,
            )
            row_data["ÁE"] = hit["AE"]
        else:
            missing += 1
            row_data["Cikkszám"] = item["key"]
            row_data["Megnevezés"] = ""
            row_data["EK"] = ""
            row_data["ÁE"] = ""

        ws.append([
            row_data.get(header, "")
            for header in headers
        ])

        row_idx = ws.max_row

        if not hit:
            for column in range(1, len(headers) + 1):
                ws.cell(row_idx, column).fill = FILL_WARN

        if "Cikkszám" in headers:
            code_col = headers.index("Cikkszám") + 1
            ws.cell(
                row_idx,
                code_col,
            ).number_format = "@"

        if "EK" in headers:
            ek_col = headers.index("EK") + 1

            if currency == "EUR":
                # Magyar Excelben 3,02 formában jelenik meg.
                ws.cell(
                    row_idx,
                    ek_col,
                ).number_format = "0.00"
            else:
                ws.cell(
                    row_idx,
                    ek_col,
                ).number_format = "0"

    resize_template_table(
        ws,
        len(headers),
    )

    out_path.parent.mkdir(
        parents=True,
        exist_ok=True,
    )
    wb.save(out_path)

    return missing, len(items)


# ============================================================
# EK IMAGE OCR
# ============================================================
def setup_tesseract() -> None:
    if (
        TESSERACT_EXE is not None
        and Path(TESSERACT_EXE).exists()
    ):
        pytesseract.pytesseract.tesseract_cmd = str(
            TESSERACT_EXE
        )


def preprocess_image(img_path: Path) -> Image.Image:
    img = Image.open(img_path).convert("L")
    img = ImageOps.autocontrast(img)
    img = img.filter(ImageFilter.SHARPEN)

    width, height = img.size
    img = img.resize(
        (width * 2, height * 2)
    )

    img = img.point(
        lambda x: 255 if x > 180 else 0
    )

    return img


def ocr_image_data(img_path: Path):
    setup_tesseract()
    img = preprocess_image(img_path)

    try:
        return pytesseract.image_to_data(
            img,
            lang="hun+eng",
            config="--psm 6",
            output_type=pytesseract.Output.DICT,
        )
    except Exception:
        return pytesseract.image_to_data(
            img,
            lang="eng",
            config="--psm 6",
            output_type=pytesseract.Output.DICT,
        )


def ocr_image_data_digits(img_path: Path):
    setup_tesseract()
    img = preprocess_image(img_path)

    return pytesseract.image_to_data(
        img,
        lang="eng",
        config=(
            "--psm 11 "
            "-c tessedit_char_whitelist=0123456789"
        ),
        output_type=pytesseract.Output.DICT,
    )


def ocr_image_data_digits_cropped(img_path: Path):
    setup_tesseract()
    img = preprocess_image(img_path)

    width, height = img.size

    left = int(width * 0.22)
    right = int(width * 0.48)
    top = int(height * 0.22)
    bottom = int(height * 0.92)

    cropped = img.crop(
        (left, top, right, bottom)
    )

    data = pytesseract.image_to_data(
        cropped,
        lang="eng",
        config=(
            "--psm 6 "
            "-c tessedit_char_whitelist=0123456789"
        ),
        output_type=pytesseract.Output.DICT,
    )

    return data, left, top


def ocr_image_text(img_path: Path) -> str:
    setup_tesseract()
    img = preprocess_image(img_path)

    try:
        return pytesseract.image_to_string(
            img,
            lang="hun+eng",
            config="--psm 6",
        )
    except Exception:
        return pytesseract.image_to_string(
            img,
            lang="eng",
            config="--psm 6",
        )


def extract_image_qty_single(
    ocr_text: str,
) -> Optional[int]:
    lines = [
        line.strip()
        for line in ocr_text.splitlines()
        if line.strip()
    ]

    for line_idx, line in enumerate(lines):
        low = line.lower()

        if "mennyis" in low:
            match = re.search(
                r"(\d{1,7})\s*(?:darab|db)?",
                line,
                flags=re.IGNORECASE,
            )

            if match:
                try:
                    return int(match.group(1))
                except ValueError:
                    pass

            if line_idx + 1 < len(lines):
                match = re.search(
                    r"(\d{1,7})\s*(?:darab|db)?",
                    lines[line_idx + 1],
                    flags=re.IGNORECASE,
                )

                if match:
                    try:
                        return int(match.group(1))
                    except ValueError:
                        pass

    for line in lines:
        low = line.lower()

        if (
            "csomagolási egység" in low
            or "vpe" in low
        ):
            match = re.search(
                r"(\d{1,7})\s*(?:darab|db)?",
                line,
                flags=re.IGNORECASE,
            )

            if match:
                try:
                    return int(match.group(1))
                except ValueError:
                    pass

    for line in reversed(lines):
        matches = list(
            re.finditer(
                r"(\d{1,7})\s*(?:darab|db)\b",
                line,
                flags=re.IGNORECASE,
            )
        )

        if matches:
            try:
                return int(matches[-1].group(1))
            except ValueError:
                pass

    return None


def extract_single_item_from_image(
    ocr_text: str,
    lookup_map: Dict[str, dict],
) -> Optional[dict]:
    lines = [
        line.strip()
        for line in ocr_text.splitlines()
        if line.strip()
    ]

    candidate_lines: List[str] = []

    for line_idx, line in enumerate(lines):
        low = line.lower()

        if "cikksz" in low:
            candidate_lines.append(line)

            if line_idx + 1 < len(lines):
                candidate_lines.append(
                    lines[line_idx + 1]
                )

    if not candidate_lines:
        candidate_lines = lines[:10]

    key = ""

    for line in candidate_lines:
        tokens = re.findall(r"\d+", line)
        key = best_lookup_candidate_from_tokens(
            tokens,
            lookup_map,
        )

        if key:
            break

    if not key:
        merged = " ".join(candidate_lines)
        tokens = re.findall(r"\d+", merged)
        key = best_lookup_candidate_from_tokens(
            tokens,
            lookup_map,
        )

    if not key:
        return None

    qty = extract_image_qty_single(ocr_text)

    return {
        "key": key,
        "qty": qty,
    }


def extract_multiple_items_from_image(
    img_path: Path,
    lookup_map: Dict[str, dict],
) -> List[dict]:
    """
    A jelenlegi, jól működő listásképes megoldás:
    teljes kép, szám-only OCR, majd lookup alapú sliding window.
    """
    data = ocr_image_data_digits(img_path)
    tokens: List[dict] = []

    for i in range(len(data["text"])):
        text = str(data["text"][i]).strip()

        if not re.fullmatch(r"\d+", text):
            continue

        try:
            confidence = float(data["conf"][i])
        except (TypeError, ValueError):
            confidence = -1

        if confidence < -1:
            continue

        tokens.append({
            "text": text,
            "left": int(data["left"][i]),
            "top": int(data["top"][i]),
        })

    if not tokens:
        return []

    tokens.sort(
        key=lambda token: (
            token["top"],
            token["left"],
        )
    )

    raw_tokens = [
        token["text"]
        for token in tokens
    ]

    found_keys: List[tuple] = []
    token_count = len(raw_tokens)

    for start_idx in range(token_count):
        accumulated = ""

        for end_idx in range(
            start_idx,
            min(start_idx + 6, token_count),
        ):
            accumulated += raw_tokens[end_idx]

            if accumulated in lookup_map:
                found_keys.append(
                    (
                        accumulated,
                        start_idx,
                        end_idx,
                    )
                )

    unique: List[tuple] = []
    seen = set()

    for key, start_idx, end_idx in sorted(
        found_keys,
        key=lambda item: (
            item[1],
            -len(item[0]),
        ),
    ):
        if key in seen:
            continue

        seen.add(key)
        unique.append(
            (
                key,
                start_idx,
                end_idx,
            )
        )

    items: List[dict] = []

    for key, _, end_idx in unique:
        qty = None

        for token_idx in range(
            end_idx + 1,
            min(end_idx + 4, token_count),
        ):
            try:
                number = int(raw_tokens[token_idx])

                if number <= 100000:
                    qty = number
                    break
            except (TypeError, ValueError):
                pass

        items.append({
            "key": key,
            "qty": qty,
        })

    return items


def process_ek_image_multi(
    img_path: Path,
    lookup_map: Dict[str, dict],
) -> None:
    items = extract_multiple_items_from_image(
        img_path,
        lookup_map,
    )

    if not items:
        ocr_text = ocr_image_text(img_path)
        single = extract_single_item_from_image(
            ocr_text,
            lookup_map,
        )

        if single:
            items = [single]

    if not items:
        print(
            f"[EK-IMG][WARN] 0 tétel: "
            f"{img_path.name}"
        )
        return

    out_path = (
        OUTPUT_DIR
        / f"EK_{img_path.stem}.xlsx"
    )

    # Képnél nincs biztosan kiolvasható vevőszám,
    # ezért alapértelmezésben HUF.
    missing, total = ek_write_output(
        items,
        out_path,
        lookup_map,
        currency=DEFAULT_CURRENCY,
    )

    print(
        f"[EK-IMG][OK] {img_path.name} -> "
        f"{out_path.name} | tételek: {total} | "
        f"nem talált: {missing} | pénznem: HUF"
    )


# ============================================================
# SAJATKESZLET POSITION-BASED PDF PARSER
# ============================================================
def extract_words_with_positions(
    pdf_path: Path,
) -> List[List[dict]]:
    pages_words: List[List[dict]] = []

    with pdfplumber.open(str(pdf_path)) as pdf:
        for page in pdf.pages:
            words = page.extract_words(
                use_text_flow=False,
                keep_blank_chars=False,
            )
            pages_words.append(words)

    return pages_words


def find_darabszam_column_x(
    words: List[dict],
):
    for word in words:
        text = (
            str(word.get("text", ""))
            .strip()
            .lower()
        )

        if text == "darabszám":
            return (
                word["x0"],
                word["x1"],
                word["top"],
            )

    return None, None, None


def find_item_rows(
    words: List[dict],
    header_top: float,
) -> List[dict]:
    grouped = defaultdict(list)

    for word in words:
        y = round(word["top"], 0)
        grouped[y].append(word)

    rows: List[dict] = []

    for y in sorted(grouped.keys()):
        if y <= header_top:
            continue

        row_words = sorted(
            grouped[y],
            key=lambda item: item["x0"],
        )

        if not row_words:
            continue

        first = str(
            row_words[0]["text"]
        ).strip()

        if not re.fullmatch(
            r"\d{1,3}",
            first,
        ):
            continue

        rest_words = row_words[1:]
        digit_words = [
            str(item["text"]).strip()
            for item in rest_words
            if re.fullmatch(
                r"\d+",
                str(item["text"]).strip(),
            )
        ]

        if len(digit_words) < 2:
            continue

        rows.append({
            "y": y,
            "words": row_words,
        })

    return rows


def extract_key_from_row_words(
    row_words: List[dict],
    lookup_map: Dict[str, dict],
) -> str:
    tokens: List[str] = []

    # Az első szó a pozíciószám.
    for word in row_words[1:]:
        text = str(word["text"]).strip()

        if re.fullmatch(r"\d+", text):
            tokens.append(text)
        else:
            break

    if not tokens:
        return ""

    return (
        best_lookup_candidate_from_tokens(
            tokens,
            lookup_map,
        )
        or ek_normalize_code(" ".join(tokens))
    )


def extract_qty_from_darabszam_column(
    words: List[dict],
    row_top: float,
    next_row_top: float,
    darab_x0: float,
) -> Optional[int]:
    candidates: List[tuple] = []

    for word in words:
        text = str(
            word.get("text", "")
        ).strip()
        x0 = word["x0"]
        top = word["top"]

        if not (
            row_top - 2
            <= top
            < next_row_top - 2
        ):
            continue

        if x0 < darab_x0 - 5:
            continue

        if re.fullmatch(r"\d{1,7}", text):
            candidates.append(
                (
                    top,
                    x0,
                    int(text),
                )
            )

    if not candidates:
        return None

    candidates.sort(
        key=lambda item: (
            item[0],
            item[1],
        )
    )

    return candidates[-1][2]


def sk_parse_items_from_pdf_positions(
    pdf_path: Path,
    lookup_map: Dict[str, dict],
) -> List[dict]:
    pages_words = extract_words_with_positions(
        pdf_path
    )
    items: List[dict] = []

    # A többoldalas PDF-ek második oldalától
    # a fejléc gyakran nincs megismételve.
    last_darab_x0 = None

    for page_idx, page_words in enumerate(
        pages_words,
        start=1,
    ):
        darab_x0, _, header_top = (
            find_darabszam_column_x(page_words)
        )

        if darab_x0 is None:
            if last_darab_x0 is None:
                continue

            darab_x0 = last_darab_x0
            header_top = -1
        else:
            last_darab_x0 = darab_x0

        rows = find_item_rows(
            page_words,
            header_top
            if header_top is not None
            else -1,
        )

        if not rows:
            continue

        for row_idx, row in enumerate(rows):
            row_top = row["y"]
            next_row_top = (
                rows[row_idx + 1]["y"]
                if row_idx + 1 < len(rows)
                else 999999
            )

            key = extract_key_from_row_words(
                row["words"],
                lookup_map,
            )

            if not key:
                continue

            qty = extract_qty_from_darabszam_column(
                page_words,
                row_top,
                next_row_top,
                darab_x0,
            )

            items.append({
                "key": key,
                "qty": qty,
            })

    return items


# =========================
# SAJATKESZLET OUTPUT
# =========================
def sk_write_output(
    items: List[dict],
    out_path: Path,
    lookup_map: Dict[str, dict],
) -> None:
    wb = load_workbook(SK_TEMPLATE_FILE)
    ws = wb.active

    if ws.max_row >= 2:
        ws.delete_rows(
            2,
            ws.max_row - 1,
        )

    headers = safe_headers_from_template(ws)

    if not headers:
        raise ValueError(
            "SAJATKESZLET_template.xlsx "
            "első sora üres."
        )

    for item in items:
        hit = lookup_map.get(item["key"])
        row_data: Dict[str, object] = {}

        if hit:
            row_data["Cikkszám"] = hit["code"]
            row_data["Megnevezés"] = hit["name"]
        else:
            row_data["Cikkszám"] = item["key"]
            row_data["Megnevezés"] = ""

        row_data["Mennyiség"] = item.get(
            "qty",
            "",
        )

        ws.append([
            row_data.get(header, "")
            for header in headers
        ])

        if "Cikkszám" in headers:
            row_idx = ws.max_row
            col_idx = (
                headers.index("Cikkszám") + 1
            )
            ws.cell(
                row_idx,
                col_idx,
            ).number_format = "@"

    resize_template_table(
        ws,
        len(headers),
    )

    out_path.parent.mkdir(
        parents=True,
        exist_ok=True,
    )
    wb.save(out_path)


# ============================================================
# UNIMAS
# ============================================================
def is_unimas_pdf_quick(text: str) -> bool:
    low = text.lower()

    return (
        "unimas" in low
        and "bestellung" in low
        and "pos artikel menge" in low
    )


def unimas_extract_order_no(
    pdf_text: str,
) -> str:
    match = re.search(
        r"\bNr\s+(\d+)\b",
        pdf_text,
        flags=re.IGNORECASE,
    )

    if not match:
        # A következő oldalakon:
        # Bestellung 626697 vom ...
        match = re.search(
            r"\bBestellung\s+(\d+)\b",
            pdf_text,
            flags=re.IGNORECASE,
        )

    return (
        match.group(1)
        if match
        else "UNIMAS_OUTPUT"
    )


def unimas_extract_items(
    pdf_text: str,
) -> List[dict]:
    """
    Példa:
      001 ADSK-01000 200Stück 0,35 70,00
    """
    items: List[dict] = []

    line_pattern = re.compile(
        r"^\s*"
        r"(\d{3})\s+"
        r"([A-Z0-9]+-[A-Z0-9]+)\s+"
        r"(\d+)\s*Stück\s+"
        r"([0-9]+(?:[.,][0-9]+)?)\s+"
        r"([0-9]+(?:[.,][0-9]+)?)"
        r"\s*$",
        flags=re.IGNORECASE,
    )

    for line in pdf_text.splitlines():
        line = line.strip()

        if not line:
            continue

        match = line_pattern.match(line)

        if not match:
            continue

        items.append({
            "Poz": match.group(1),
            "Unimas cikkszám": (
                match.group(2).upper()
            ),
            "Menge": int(match.group(3)),
            "Einzelpreis EUR %": (
                parse_float_any(match.group(4))
            ),
            "Gesamtpreis EUR": (
                parse_float_any(match.group(5))
            ),
        })

    return items


def unimas_write_output(
    rows: List[dict],
    out_path: Path,
) -> None:
    wb = load_workbook(UNIMAS_TEMPLATE_FILE)
    ws = wb.active

    if ws.max_row >= 2:
        ws.delete_rows(
            2,
            ws.max_row - 1,
        )

    headers = safe_headers_from_template(ws)

    if not headers:
        raise ValueError(
            "UNIMAS_template.xlsx első sora üres."
        )

    for row in rows:
        row_data = {
            "Unimas cikkszám": row.get(
                "Unimas cikkszám",
                "",
            ),
            "Menge": row.get(
                "Menge",
                "",
            ),
            "Einzelpreis EUR %": row.get(
                "Einzelpreis EUR %",
                "",
            ),
            "Gesamtpreis EUR": row.get(
                "Gesamtpreis EUR",
                "",
            ),
        }

        ws.append([
            row_data.get(header, "")
            for header in headers
        ])

        current_row = ws.max_row

        if "Unimas cikkszám" in headers:
            code_col = (
                headers.index(
                    "Unimas cikkszám"
                )
                + 1
            )
            ws.cell(
                current_row,
                code_col,
            ).number_format = "@"

        for price_header in (
            "Einzelpreis EUR %",
            "Gesamtpreis EUR",
        ):
            if price_header in headers:
                price_col = (
                    headers.index(price_header) + 1
                )
                ws.cell(
                    current_row,
                    price_col,
                ).number_format = "0.00"

    resize_template_table(
        ws,
        len(headers),
    )

    out_path.parent.mkdir(
        parents=True,
        exist_ok=True,
    )
    wb.save(out_path)


def process_pdf_unimas(
    pdf_path: Path,
    full_text: str,
) -> None:
    rows = unimas_extract_items(full_text)

    if not rows:
        print(
            f"[UNIMAS][WARN] 0 tétel: "
            f"{pdf_path.name}"
        )
        return

    order_no = unimas_extract_order_no(
        full_text
    )

    out_path = (
        OUTPUT_DIR
        / f"{order_no}.xlsx"
    )

    unimas_write_output(
        rows,
        out_path,
    )

    print(
        f"[UNIMAS][OK] {pdf_path.name} -> "
        f"{out_path.name} | sorok: {len(rows)}"
    )


# ============================================================
# GE
# ============================================================
GE_LOOKUP_SHEET_NAME = "Sheet2"
GE_LOOKUP_COL_ITEM = "GE cikkszám"
GE_LOOKUP_COL_RECA = "RECA cikkszám"


def ge_load_lookup_map() -> Dict[str, str]:
    _, ge_map = _load_lookup_maps_with_cache()
    return ge_map

def ge_extract_release_number(
    pdf_text: str,
) -> str:
    match = re.search(
        r"\b\d{6,}-\d{1,}\b",
        pdf_text,
    )

    return (
        match.group(0)
        if match
        else "GE_OUTPUT"
    )


def ge_extract_items_from_text(
    full_text: str,
) -> List[dict]:
    rows: List[dict] = []

    price_pattern = (
        r"(?:\d+(?:[.,]\d+)?|[.,]\d+)"
    )

    item_pattern = re.compile(
        r"\bItem:\s*([0-9A-Za-z\-]+)"
    )

    each_pattern = re.compile(
        rf"\bEACH\s+"
        rf"(\d+)\s+"
        rf"({price_pattern})\s+"
        rf"({price_pattern})"
    )

    item_matches = list(
        item_pattern.finditer(full_text)
    )

    for item_idx, match in enumerate(
        item_matches
    ):
        item_code = match.group(1).strip()
        start = match.end()
        end = (
            item_matches[item_idx + 1].start()
            if item_idx + 1 < len(item_matches)
            else len(full_text)
        )
        block = full_text[start:end]

        each_match = each_pattern.search(block)

        if not each_match:
            continue

        qty = int(each_match.group(1))
        price = parse_float_any(
            each_match.group(2)
        )

        rows.append({
            "GE cikkszám": item_code,
            "Mennyiség": qty,
            "Nettó egységár": price,
        })

    return rows


def ge_write_output(
    rows: List[dict],
    out_path: Path,
    ge_lookup_map: Dict[str, str],
) -> None:
    wb = load_workbook(GE_TEMPLATE_FILE)
    ws = wb.active

    if ws.max_row >= 2:
        ws.delete_rows(
            2,
            ws.max_row - 1,
        )

    headers = safe_headers_from_template(ws)

    if not headers:
        raise ValueError(
            "GE_template.xlsx első sora üres."
        )

    for row in rows:
        ge_code = row.get(
            "GE cikkszám",
            "",
        )
        reca = ge_lookup_map.get(
            norm_key(ge_code),
            "",
        )

        row_data = {
            "GE Cikkszám": ge_code,
            "Reca cikkszám": reca,
            "Mennyiség": row.get(
                "Mennyiség",
                "",
            ),
            "GE ár": row.get(
                "Nettó egységár",
                "",
            ),
        }

        ws.append([
            row_data.get(header, "")
            for header in headers
        ])

        if "Reca cikkszám" in headers:
            current_row = ws.max_row
            col_idx = (
                headers.index(
                    "Reca cikkszám"
                )
                + 1
            )
            ws.cell(
                current_row,
                col_idx,
            ).number_format = "@"

    resize_template_table(
        ws,
        len(headers),
    )

    out_path.parent.mkdir(
        parents=True,
        exist_ok=True,
    )
    wb.save(out_path)


# ============================================================
# TYPE DETECT
# ============================================================
def is_ge_pdf_quick(
    first_pages_text: str,
) -> bool:
    low = first_pages_text.lower()
    score = 0

    if "item:" in low:
        score += 1

    if "each" in low:
        score += 1

    if "net unit price" in low:
        score += 1

    if "blanket release" in low:
        score += 1

    return score >= 2


# ============================================================
# PROCESSORS
# ============================================================
def process_pdf_ek(
    pdf_path: Path,
    lookup_map: Dict[str, dict],
    customer_currency_map: Optional[
        Dict[str, str]
    ] = None,
) -> None:
    """
    A customer_currency_map opcionális, hogy a régi GUI-hívások
    (process_pdf_ek(path, lookup_map)) továbbra is működjenek.
    """
    full_text = extract_text_from_pdf_all(
        pdf_path
    )
    customer_no = ek_extract_customer_no(
        full_text
    )

    if customer_currency_map is None:
        customer_currency_map = (
            load_customer_currency_map()
        )

    currency = get_customer_currency(
        customer_no,
        customer_currency_map,
    )

    items = ek_parse_pdf_items(
        full_text,
        lookup_map,
    )

    if not items:
        print(
            f"[EK][WARN] 0 tétel: "
            f"{pdf_path.name}"
        )
        return

    out_path = (
        OUTPUT_DIR
        / f"{customer_no} EK.xlsx"
    )

    missing, total = ek_write_output(
        items,
        out_path,
        lookup_map,
        currency=currency,
    )

    print(
        f"[EK][OK] {pdf_path.name} -> "
        f"{out_path.name} | vevő: {customer_no} | "
        f"pénznem: {currency} | tételek: {total} | "
        f"nem talált: {missing}"
    )


def process_pdf_sk(
    pdf_path: Path,
    lookup_map: Dict[str, dict],
) -> None:
    items = sk_parse_items_from_pdf_positions(
        pdf_path,
        lookup_map,
    )

    if not items:
        print(
            f"[SK][WARN] 0 tétel: "
            f"{pdf_path.name}"
        )
        return

    out_path = (
        OUTPUT_DIR
        / f"SAJATKESZLET_{pdf_path.stem}.xlsx"
    )

    sk_write_output(
        items,
        out_path,
        lookup_map,
    )

    print(
        f"[SK][OK] {pdf_path.name} -> "
        f"{out_path.name} | sorok: {len(items)}"
    )


def process_pdf_ge(
    pdf_path: Path,
    full_text: str,
    ge_lookup_map: Dict[str, str],
) -> None:
    rows = ge_extract_items_from_text(
        full_text
    )

    if not rows:
        print(
            f"[GE][WARN] 0 tétel: "
            f"{pdf_path.name}"
        )
        return

    release_no = ge_extract_release_number(
        full_text
    )
    out_path = (
        OUTPUT_DIR
        / f"{release_no}.xlsx"
    )

    ge_write_output(
        rows,
        out_path,
        ge_lookup_map,
    )

    print(
        f"[GE][OK] {pdf_path.name} -> "
        f"{out_path.name} | sorok: {len(rows)}"
    )


# ============================================================
# WATCHDOG
# ============================================================
class EkHandler(FileSystemEventHandler):
    def __init__(
        self,
        lookup_map: Dict[str, dict],
        customer_currency_map: Optional[
            Dict[str, str]
        ] = None,
    ):
        self.lookup_map = lookup_map
        self.customer_currency_map = (
            customer_currency_map or {}
        )

    def on_created(self, event):
        if event.is_directory:
            return

        path = Path(event.src_path)
        suffix = path.suffix.lower()

        if suffix not in {
            ".pdf",
            ".jpg",
            ".jpeg",
            ".png",
            ".webp",
        }:
            return

        wait_until_file_ready(path)

        try:
            if suffix == ".pdf":
                process_pdf_ek(
                    path,
                    self.lookup_map,
                    self.customer_currency_map,
                )
            else:
                process_ek_image_multi(
                    path,
                    self.lookup_map,
                )
        except Exception as error:
            print(
                f"[EK][ERROR] {path.name}: "
                f"{error}"
            )


class AutoHandler(FileSystemEventHandler):
    def __init__(
        self,
        lookup_map: Dict[str, dict],
        ge_lookup_map: Dict[str, str],
    ):
        self.lookup_map = lookup_map
        self.ge_lookup_map = ge_lookup_map

    def on_created(self, event):
        if event.is_directory:
            return

        path = Path(event.src_path)

        if path.suffix.lower() != ".pdf":
            return

        wait_until_file_ready(path)

        try:
            head_text = extract_text_from_pdf_pages(
                path,
                max_pages=2,
            )
            full_text = extract_text_from_pdf_all(
                path
            )

            if is_unimas_pdf_quick(head_text):
                process_pdf_unimas(
                    path,
                    full_text,
                )
            elif is_ge_pdf_quick(head_text):
                process_pdf_ge(
                    path,
                    full_text,
                    self.ge_lookup_map,
                )
            else:
                process_pdf_sk(
                    path,
                    self.lookup_map,
                )

        except Exception as error:
            print(
                f"[AUTO][ERROR] {path.name}: "
                f"{error}"
            )


# ============================================================
# MAIN
# ============================================================
def main():
    INPUT_DIR.mkdir(
        parents=True,
        exist_ok=True,
    )
    INPUT_EK_DIR.mkdir(
        parents=True,
        exist_ok=True,
    )
    OUTPUT_DIR.mkdir(
        parents=True,
        exist_ok=True,
    )

    required_templates = [
        EK_TEMPLATE_FILE,
        GE_TEMPLATE_FILE,
        SK_TEMPLATE_FILE,
        UNIMAS_TEMPLATE_FILE,
    ]

    for template in required_templates:
        if not template.exists():
            print(
                f"[HIBA] Hiányzik template: "
                f"{template}"
            )
            return

    if not LOOKUP_FILE.exists():
        print(
            f"[HIBA] Hiányzik lookup: "
            f"{LOOKUP_FILE}"
        )
        return

    print(f"[INFO] ENGINE VERZIÓ: {ENGINE_VERSION}", flush=True)
    print("[INFO] Tartós cache-ek ellenőrzése...", flush=True)

    ek_lookup_map = ek_load_lookup_map()
    _ = build_lookup_lengths(ek_lookup_map)
    ge_lookup_map = ge_load_lookup_map()
    customer_currency_map = (
        load_customer_currency_map()
    )

    print(
        f"[INFO] EK/SK lookup: "
        f"{len(ek_lookup_map)} | "
        f"GE lookup: {len(ge_lookup_map)} | "
        f"vevői pénznemek: "
        f"{len(customer_currency_map)}"
    )

    print("[INFO] Minden Excel/cache beolvasva.", flush=True)
    print("[INFO] A program készen áll a fájlok fogadására.", flush=True)

    observer = Observer()

    observer.schedule(
        EkHandler(
            ek_lookup_map,
            customer_currency_map,
        ),
        str(INPUT_EK_DIR),
        recursive=False,
    )

    observer.schedule(
        AutoHandler(
            ek_lookup_map,
            ge_lookup_map,
        ),
        str(INPUT_DIR),
        recursive=False,
    )

    observer.start()

    print(
        f"Watching EK:   {INPUT_EK_DIR} "
        f"(PDF + kép)"
    )
    print(
        f"Watching AUTO: {INPUT_DIR} "
        f"(GE vs SAJATKESZLET vs UNIMAS)"
    )
    print(f"Output:        {OUTPUT_DIR}")
    print(
        f"EUR árfolyam:  1 EUR = "
        f"{EUR_RATE:g} HUF"
    )
    print("Kilépés: Ctrl+C")

    try:
        while True:
            time.sleep(1)
    except KeyboardInterrupt:
        observer.stop()

    observer.join()


if __name__ == "__main__":
    main()
